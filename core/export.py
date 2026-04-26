"""
Excel + ZIP export builder.

Produces a ZIP archive containing:
  - overview.xlsx — 4 sheets (Summary, Matched, Unmatched, Unused)
  - receipts/<yyyy-mm>/<renamed-file> — copies of matched receipts,
    renamed to `<date>_<amount>_<description>.<ext>`
  - receipts/_unused/<original-name> — receipts that weren't assigned

The Excel file is structured so the accountant can:
  - Immediately see what needs follow-up (Unmatched sheet).
  - Drag renamed receipts into their bookkeeping software as evidence.
  - Spot duplicates or wrong-account receipts (Unused sheet).
"""

from __future__ import annotations

import os
import re
import zipfile
from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .sessions import SessionStore


class ExportBuilder:
    """Builds the accountant-ready ZIP+XLSX bundle for a session."""

    # openpyxl styles (class-level so they're built once)
    _HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    _HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
    _HEADER_ALIGN = Alignment(horizontal="left", vertical="center")
    _CONF_FILLS = {
        "High":   PatternFill("solid", fgColor="DCFCE7"),
        "Medium": PatternFill("solid", fgColor="FEF3C7"),
        "Low":    PatternFill("solid", fgColor="FFEDD5"),
        "Manual": PatternFill("solid", fgColor="F3E8FF"),
    }
    _THIN_BORDER = Border(bottom=Side(style="thin", color="E2E8F0"))

    def __init__(self, store: SessionStore):
        self.store = store

    def build(self, session_id: str, results: list[dict]) -> tuple[bytes, str]:
        """Returns (zip_bytes, download_filename)."""
        session_dir = self.store.root / session_id
        if not session_dir.exists():
            raise ValueError("Session not found or expired.")

        matched, unmatched = self._partition_results(results)
        assigned_filenames = {r["matchedReceipt"]["fileName"] for r in matched}
        unused_filenames = self._find_unused(session_dir, assigned_filenames)

        # Build rename mapping: for each matched receipt → new monthly path
        rename_map = self._build_rename_map(matched)

        # Generate Excel file
        xlsx_bytes = self._build_excel(matched, unmatched, unused_filenames, rename_map)

        # Pack everything into a ZIP
        zip_bytes = self._build_zip(session_dir, xlsx_bytes, rename_map, unused_filenames)
        filename = self._download_filename(matched + unmatched)
        return zip_bytes, filename

    # ── Categorisation ────────────────────────────────────────────────────

    @staticmethod
    def _partition_results(results: list[dict]) -> tuple[list[dict], list[dict]]:
        matched, unmatched = [], []
        for r in results:
            (matched if r.get("matchedReceipt") else unmatched).append(r)
        return matched, unmatched

    def _find_unused(self, session_dir: Path, assigned: set[str]) -> list[str]:
        return sorted(
            f.name for f in session_dir.iterdir()
            if f.is_file() and f.name not in assigned
        )

    # ── Filename generation ───────────────────────────────────────────────

    def _build_rename_map(self, matched: list[dict]) -> dict[str, str]:
        """Maps original filename → new relative path (e.g. '2026-02/2026-02-19_380.80_GMAP.pdf')."""
        rename_map: dict[str, str] = {}
        used_paths: set[str] = set()

        for r in matched:
            original = r["matchedReceipt"]["fileName"]
            txn = r["transaction"]
            month = txn["date"][:7]  # YYYY-MM
            ext = Path(original).suffix or ""

            base_name = self._make_safe_filename(txn, ext)
            rel_path = f"{month}/{base_name}"
            rel_path = self._ensure_unique(rel_path, used_paths)
            used_paths.add(rel_path)
            rename_map[original] = rel_path

        return rename_map

    @staticmethod
    def _make_safe_filename(txn: dict, extension: str) -> str:
        """Builds `YYYY-MM-DD_AMOUNT_DESCRIPTION.ext` with sanitised parts."""
        date = txn["date"]
        amount = f"{txn['amount']:.2f}"
        raw_desc = txn.get("description", "")

        # Keep unicode letters/digits/hyphens; drop bullet chars, slashes, etc.
        safe = re.sub(r"[^\w\s-]", "", raw_desc, flags=re.UNICODE)
        safe = re.sub(r"\s+", "-", safe.strip())
        safe = safe[:40].rstrip("-") or "transaction"

        return f"{date}_{amount}_{safe}{extension}"

    @staticmethod
    def _ensure_unique(rel_path: str, used: set[str]) -> str:
        if rel_path not in used:
            return rel_path
        stem, ext = os.path.splitext(rel_path)
        counter = 2
        while f"{stem}_{counter}{ext}" in used:
            counter += 1
        return f"{stem}_{counter}{ext}"

    # ── Excel generation ──────────────────────────────────────────────────

    def _build_excel(self, matched: list[dict], unmatched: list[dict],
                     unused: list[str], rename_map: dict[str, str]) -> bytes:
        wb = openpyxl.Workbook()

        self._write_summary_sheet(wb.active, matched, unmatched, unused)
        self._write_matched_sheet(wb.create_sheet("Matched"), matched, rename_map)
        self._write_unmatched_sheet(wb.create_sheet("Unmatched"), unmatched)
        self._write_unused_sheet(wb.create_sheet("Unused Receipts"), unused)

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _write_summary_sheet(self, ws, matched: list[dict], unmatched: list[dict],
                             unused: list[str]) -> None:
        ws.title = "Summary"
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 20

        # Period derived from transaction dates
        all_results = matched + unmatched
        dates = [r["transaction"]["date"] for r in all_results if r["transaction"].get("date")]
        period = f"{min(dates)} to {max(dates)}" if dates else "—"

        # Counts by confidence and type
        conf_counts = {"High": 0, "Medium": 0, "Low": 0, "Manual": 0, "None": 0}
        total_debit = total_credit = 0.0
        for r in all_results:
            conf_counts[r.get("confidence", "None")] = conf_counts.get(r.get("confidence", "None"), 0) + 1
            t = r["transaction"]
            if t.get("type") == "Credit":
                total_credit += t["amount"]
            else:
                total_debit += t["amount"]

        # Write sections
        self._write_heading(ws, 1, "DocMatcher Export")
        ws.cell(row=2, column=1, value="Period").font = Font(bold=True)
        ws.cell(row=2, column=2, value=period)
        ws.cell(row=3, column=1, value="Generated").font = Font(bold=True)
        ws.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

        self._write_heading(ws, 5, "Transactions")
        rows = [
            ("Total", len(all_results)),
            ("Matched – high confidence", conf_counts["High"]),
            ("Matched – medium confidence", conf_counts["Medium"]),
            ("Matched – low confidence", conf_counts["Low"]),
            ("Manually assigned", conf_counts["Manual"]),
            ("Unmatched (follow up)", conf_counts["None"]),
        ]
        for i, (label, value) in enumerate(rows, start=6):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=value)

        self._write_heading(ws, 13, "Financial summary")
        ws.cell(row=14, column=1, value="Total expenses")
        c = ws.cell(row=14, column=2, value=total_debit)
        c.number_format = '#,##0.00 "EUR"'
        ws.cell(row=15, column=1, value="Total income")
        c = ws.cell(row=15, column=2, value=total_credit)
        c.number_format = '#,##0.00 "EUR"'

        self._write_heading(ws, 17, "Receipts")
        ws.cell(row=18, column=1, value="Matched to transactions").value = "Matched to transactions"
        ws.cell(row=18, column=2, value=len(matched))
        ws.cell(row=19, column=1, value="Unused (review for duplicates)")
        ws.cell(row=19, column=2, value=len(unused))

    def _write_matched_sheet(self, ws, matched: list[dict],
                             rename_map: dict[str, str]) -> None:
        headers = ["Date", "Description", "Amount", "Currency", "Type",
                   "Receipt File", "Confidence", "OCR Amount", "OCR Date"]
        self._write_header_row(ws, headers)

        for i, r in enumerate(matched, start=2):
            t = r["transaction"]
            mr = r["matchedReceipt"]
            row_data = [
                t["date"],
                t["description"],
                t["amount"],
                t["currency"],
                t["type"],
                rename_map.get(mr["fileName"], mr["fileName"]),
                r["confidence"],
                mr.get("extractedAmount"),
                mr.get("extractedDate") or "",
            ]
            for j, val in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                cell.border = self._THIN_BORDER
                if j in (3, 8) and val is not None and val != "":
                    cell.number_format = "#,##0.00"

            # Highlight confidence cell
            fill = self._CONF_FILLS.get(r["confidence"])
            if fill:
                ws.cell(row=i, column=7).fill = fill

        self._freeze_and_size(ws, [12, 40, 12, 10, 10, 42, 13, 12, 12])

    def _write_unmatched_sheet(self, ws, unmatched: list[dict]) -> None:
        headers = ["Date", "Description", "Amount", "Currency", "Type", "Notes"]
        self._write_header_row(ws, headers)

        for i, r in enumerate(unmatched, start=2):
            t = r["transaction"]
            row_data = [t["date"], t["description"], t["amount"],
                        t["currency"], t["type"], ""]
            for j, val in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                cell.border = self._THIN_BORDER
                if j == 3:
                    cell.number_format = "#,##0.00"

        self._freeze_and_size(ws, [12, 40, 12, 10, 10, 30])

    def _write_unused_sheet(self, ws, unused_filenames: list[str]) -> None:
        headers = ["File Name", "Notes"]
        self._write_header_row(ws, headers)

        for i, fn in enumerate(unused_filenames, start=2):
            ws.cell(row=i, column=1, value=fn).border = self._THIN_BORDER
            ws.cell(row=i, column=2, value="").border = self._THIN_BORDER

        self._freeze_and_size(ws, [50, 30])

    def _write_heading(self, ws, row: int, text: str) -> None:
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=12, color="1E3A8A")

    def _write_header_row(self, ws, headers: list[str]) -> None:
        for col, name in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = self._HEADER_FONT
            cell.fill = self._HEADER_FILL
            cell.alignment = self._HEADER_ALIGN

    def _freeze_and_size(self, ws, widths: list[int]) -> None:
        ws.freeze_panes = "A2"
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 22

    # ── ZIP packaging ─────────────────────────────────────────────────────

    def _build_zip(self, session_dir: Path, xlsx_bytes: bytes,
                   rename_map: dict[str, str], unused: list[str]) -> bytes:
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("overview.xlsx", xlsx_bytes)

            for original, new_path in rename_map.items():
                src = session_dir / original
                if src.exists():
                    zf.write(src, f"receipts/{new_path}")

            for fn in unused:
                src = session_dir / fn
                if src.exists():
                    zf.write(src, f"receipts/_unused/{fn}")

        return zip_buffer.getvalue()

    @staticmethod
    def _download_filename(results: list[dict]) -> str:
        dates = [r["transaction"]["date"] for r in results
                 if r["transaction"].get("date")]
        if not dates:
            return "docmatcher-export.zip"
        earliest = min(dates)[:7]  # YYYY-MM
        latest = max(dates)[:7]
        suffix = earliest if earliest == latest else f"{earliest}_to_{latest}"
        return f"docmatcher-export_{suffix}.zip"
