"""
DocMatcher – Tax Document Reconciliation
=========================================

A small Flask server that matches bank transactions (CSV) to receipt files
(JPG, PNG, TIFF, PDF) using Tesseract OCR. Designed for tax accountants
reconciling bookkeeping records.

Run:
    python server.py

Then open http://localhost:7860 in your browser.
"""

from __future__ import annotations

import io
import json
import os
import re
import shutil
import time
import uuid
import zipfile
import logging
import traceback
from dataclasses import dataclass, asdict
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

import openpyxl
import pandas as pd
import pdfplumber
import pytesseract
from flask import Flask, Response, abort, jsonify, request, send_file, send_from_directory
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image
from werkzeug.utils import secure_filename


# ── Config ────────────────────────────────────────────────────────────────

PORT = int(os.environ.get("PORT", 7860))

# Where uploaded receipts are temporarily stored so the UI can preview them.
# Each match request creates a new session folder under UPLOAD_DIR.
UPLOAD_DIR = Path(os.environ.get("UPLOAD_DIR", "uploads")).resolve()
SESSION_TTL_HOURS = int(os.environ.get("SESSION_TTL_HOURS", 24))

# Amount matching tolerances
AMOUNT_TOLERANCE_PERCENT = 0.01   # 1 %
AMOUNT_MIN_TOLERANCE = 0.02       # €0.02 absolute minimum
DAYS_TOLERANCE_HIGH = 3
DAYS_TOLERANCE_MEDIUM = 14
MAX_ALTERNATIVES = 3

# Tesseract language(s) – add 'deu' or 'ell' once the trained data is installed
OCR_LANGUAGES = os.environ.get("OCR_LANGS", "eng")

# CSV column name aliases (auto-detected).
# Order matters: first match wins, so put more-specific names first.
DATE_ALIASES = [
    # Revolut Business — prefer completion date over started date
    "date completed (utc)", "date completed",
    "date started (utc)",   "date started",
    # Other common bank formats
    "value date", "booking date", "transaction date", "posting date",
    "valuedate",  "bookingdate",  "transactiondate",
    "date", "datum",
]
DESCRIPTION_ALIASES = [
    "description", "details", "narrative", "memo",
    "particulars", "remarks", "merchant", "payee",
]
REFERENCE_ALIASES = ["reference", "ref", "note", "notes"]

AMOUNT_ALIASES   = ["amount", "value", "net amount"]
DEBIT_ALIASES    = ["debit", "withdrawal", "payments"]
CREDIT_ALIASES   = ["credit", "deposit", "receipts"]
CURRENCY_ALIASES = ["payment currency", "currency", "ccy"]
STATE_ALIASES    = ["state", "status"]

# Only these transaction states are considered final and matched.
# Used when a CSV has a state/status column (e.g. Revolut).
FINAL_STATES = {"COMPLETED", "DONE", "POSTED", "BOOKED", "SETTLED"}

# Logging
logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s  %(levelname)-5s  %(message)s",
                    datefmt="%H:%M:%S")
log = logging.getLogger("docmatcher")


# ── Domain Models ─────────────────────────────────────────────────────────

@dataclass(frozen=True)
class BankTransaction:
    id: str
    date: str           # ISO yyyy-mm-dd
    description: str
    amount: float
    currency: str
    type: str           # 'Debit' | 'Credit'


@dataclass(frozen=True)
class Receipt:
    fileName: str
    extractedAmount: Optional[float]
    extractedDate: Optional[str]


# ── Parsing Helpers ───────────────────────────────────────────────────────

def parse_number(raw: Optional[str]) -> Optional[float]:
    """Parse a number string, handling EU (1.234,56) and Anglo (1,234.56) formats."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    s = str(raw).replace("€", "").replace("$", "").replace("£", "").strip()
    if not s:
        return None

    last_dot = s.rfind(".")
    last_comma = s.rfind(",")
    if last_comma > last_dot:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", "")

    try:
        return float(s)
    except ValueError:
        return None


def parse_date(raw: Optional[str]) -> Optional[str]:
    """Parse a date into ISO yyyy-mm-dd, trying common European formats."""
    if raw is None:
        return None
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y",
                "%d.%m.%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


# Month names in German and English (lower-case, no umlaut variations both listed)
MONTH_NAMES = {
    # German
    "januar": 1, "februar": 2, "märz": 3, "marz": 3, "april": 4,
    "mai": 5, "juni": 6, "juli": 7, "august": 8,
    "september": 9, "oktober": 10, "november": 11, "dezember": 12,
    # English
    "january": 1, "march": 3, "may": 5, "june": 6, "july": 7, "october": 10,
    # Common English abbreviations
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "jun": 6, "jul": 7,
    "aug": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
}


def parse_date_with_month_name(text: str) -> Optional[str]:
    """Extracts the first 'day month-name year' date from text, e.g. '19. Februar 2026'."""
    match = re.search(r"\b(\d{1,2})\.?\s+([A-Za-zäöüÄÖÜ]+)\s+(\d{4})\b", text)
    if not match:
        return None
    day, month_name, year = match.groups()
    month = MONTH_NAMES.get(month_name.lower())
    if not month:
        return None
    try:
        return f"{int(year):04d}-{month:02d}-{int(day):02d}"
    except ValueError:
        return None


def _find_column_index(headers_lower: list[str], aliases: list[str]) -> int:
    for alias in aliases:
        if alias in headers_lower:
            return headers_lower.index(alias)
    return -1


# ── CSV Parser ────────────────────────────────────────────────────────────

class CsvParser:
    """
    Parses bank-statement CSV files into BankTransaction objects.
    Auto-detects column layout across common export formats (Revolut Business,
    Bank of Cyprus, Hellenic Bank, generic debit/credit exports).
    """

    def parse(self, content: bytes) -> list[BankTransaction]:
        text = content.decode("utf-8-sig", errors="replace")
        separator = ";" if text.count(";") > text.count(",") else ","

        df = pd.read_csv(io.StringIO(text), sep=separator,
                         dtype=str, on_bad_lines="skip")
        df.columns = [c.strip() for c in df.columns]
        headers_lower = [c.lower() for c in df.columns]

        cols = self._resolve_columns(headers_lower)
        if cols["date"] == -1:
            log.warning("No date column found. Headers: %s", df.columns.tolist())
            return []

        transactions: list[BankTransaction] = []
        skipped_states: dict[str, int] = {}

        for idx, row in df.iterrows():
            cells = row.tolist()

            # Skip non-final transactions (e.g. PENDING, DECLINED, REVERTED)
            state = self._get_cell(cells, cols["state"]).upper()
            if cols["state"] != -1 and state and state not in FINAL_STATES:
                skipped_states[state] = skipped_states.get(state, 0) + 1
                continue

            txn = self._parse_row(cells, idx, cols)
            if txn is not None:
                transactions.append(txn)

        if skipped_states:
            log.info("Skipped non-final transactions: %s", skipped_states)

        return transactions

    # ── Column Resolution ─────────────────────────────────────────────────

    def _resolve_columns(self, headers_lower: list[str]) -> dict[str, int]:
        return {
            "date":        _find_column_index(headers_lower, DATE_ALIASES),
            "description": _find_column_index(headers_lower, DESCRIPTION_ALIASES),
            "reference":   _find_column_index(headers_lower, REFERENCE_ALIASES),
            "amount":      _find_column_index(headers_lower, AMOUNT_ALIASES),
            "debit":       _find_column_index(headers_lower, DEBIT_ALIASES),
            "credit":      _find_column_index(headers_lower, CREDIT_ALIASES),
            "currency":    _find_column_index(headers_lower, CURRENCY_ALIASES),
            "state":       _find_column_index(headers_lower, STATE_ALIASES),
        }

    # ── Row Parsing ───────────────────────────────────────────────────────

    def _parse_row(self, cells: list, idx: int, cols: dict[str, int]
                   ) -> Optional[BankTransaction]:

        date = parse_date(self._get_cell(cells, cols["date"]))
        if date is None:
            return None

        amount, txn_type = self._resolve_amount(
            cells, cols["amount"], cols["debit"], cols["credit"]
        )
        if amount == 0:
            return None

        currency = self._get_cell(cells, cols["currency"]) or "EUR"
        description = self._build_description(
            self._get_cell(cells, cols["description"]),
            self._get_cell(cells, cols["reference"]),
        )

        return BankTransaction(
            id=f"TXN-{idx + 1:06d}",
            date=date,
            description=description,
            amount=round(amount, 2),
            currency=currency,
            type=txn_type,
        )

    @staticmethod
    def _get_cell(cells: list, col_index: int) -> str:
        if col_index < 0 or col_index >= len(cells):
            return ""
        raw = cells[col_index]
        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
            return ""
        return str(raw).strip()

    @staticmethod
    def _build_description(description: str, reference: str) -> str:
        """Combines description and reference fields intelligently."""
        description = description.strip()
        reference = reference.strip()

        # Skip references that are just technical IDs (all digits, UUIDs, etc.)
        if reference and re.fullmatch(r"[\d\-/]+", reference):
            reference = ""

        if description and reference and reference.lower() not in description.lower():
            return f"{description} · {reference}"
        if description:
            return description
        if reference:
            return reference
        return "—"

    def _resolve_amount(self, cells: list,
                        col_amount: int, col_debit: int, col_credit: int
                        ) -> tuple[float, str]:
        """Returns (absolute_amount, type) where type is 'Debit' or 'Credit'."""
        if col_amount != -1:
            value = parse_number(self._get_cell(cells, col_amount))
            if value is not None and value != 0:
                return abs(value), ("Debit" if value < 0 else "Credit")

        debit = parse_number(self._get_cell(cells, col_debit))
        credit = parse_number(self._get_cell(cells, col_credit))

        if debit and debit > 0:
            return debit, "Debit"
        if credit and credit > 0:
            return credit, "Credit"
        return 0, "Debit"


# ── OCR Service ───────────────────────────────────────────────────────────

class OcrService:
    """Extracts amounts and dates from receipt files via Tesseract / pdfplumber."""

    # ── Amount patterns ────────────────────────────────────────────────
    # Matches a number with decimals, supporting both EU (1.234,56) and Anglo (1,234.56) formats.
    AMOUNT_NUMBER = r"[\d]{1,6}(?:[.,]\d{3})*[.,]\d{2}"

    # Priority 1 — "paid amount" keywords. Highest confidence — this is what hit the bank.
    # OCR-tolerant: Tesseract often reads "Bezahlter" as "Bezahiter" (l → i substitution).
    # The `\w{2,6}` catches any of: Bezahlter, Bezahiter, bezahlten, etc.
    PAID_KEYWORD_PATTERN = re.compile(
        r"(?:"
        r"bezah\w{2,6}\s+betrag"       # Bezahlter Betrag (and OCR variants)
        r"|zahlbetrag"
        r"|(?<![a-z])paid(?:\s+amount)?"
        r"|amount\s+(?:paid|charged)"
        r"|payment\s+(?:amount|received)"
        r"|total\s+(?:paid|charged)"
        r"|charged\s+(?:to|amount)"
        r")"
        r"\s*(?:\([^)]*\))?"           # optional parenthetical, e.g. "(EUR)"
        r"\s*[:\s]*"
        r"(?:EUR|€|USD|\$|GBP|£)?\s*"
        rf"({AMOUNT_NUMBER})",
        re.IGNORECASE,
    )

    # Priority 2 — "total" keywords. Covers English, German, Greek.
    TOTAL_KEYWORD_PATTERN = re.compile(
        r"(?:"
        r"(?<![a-z])total\s*(?:amount|due|payable)?"
        r"|grand\s*total"
        r"|amount\s*due"
        r"|balance\s*due"
        r"|gesamt(?:betrag)?"
        r"|endbetrag"
        r"|rechnungsbetrag"
        r"|(?<![a-zäöü])summe"
        r"|σύνολο"
        r")\b\s*[:\s]*"
        r"(?:EUR|€|USD|\$|GBP|£)?\s*"
        rf"({AMOUNT_NUMBER})",
        re.IGNORECASE,
    )

    # Priority 3 — any currency-marked amount (fallback, e.g. for 2-column layouts)
    CURRENCY_AMOUNT_PATTERN = re.compile(
        rf"(?:(?:EUR|€|USD|\$|GBP|£)\s*({AMOUNT_NUMBER})"
        rf"|({AMOUNT_NUMBER})\s*(?:EUR|€|USD|\$|GBP|£))",
        re.IGNORECASE,
    )

    # Keywords that mark an amount as HISTORICAL / superseded and should be ignored.
    # E.g. on Airbnb receipts: "Vorheriger Gesamtbetrag" (previous total before adjustment).
    STALE_AMOUNT_PREFIXES = ("vorherig", "previous", "alt", "old", "original")

    # ── Date patterns ──────────────────────────────────────────────────
    DATE_PATTERNS = [
        re.compile(r"\b(\d{4}-\d{2}-\d{2})\b"),
        re.compile(r"\b(\d{2}[./]\d{2}[./]\d{4})\b"),
    ]

    IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".tiff", ".tif", ".bmp"}

    AMOUNT_MIN = 0.5
    AMOUNT_MAX = 999_999

    def extract(self, content: bytes, filename: str) -> Receipt:
        ext = os.path.splitext(filename)[1].lower()
        text = self._extract_text(content, ext)

        return Receipt(
            fileName=filename,
            extractedAmount=self._extract_amount(text),
            extractedDate=self._extract_date(text),
        )

    # ── Text extraction ────────────────────────────────────────────────

    def _extract_text(self, content: bytes, extension: str) -> str:
        try:
            if extension == ".pdf":
                with pdfplumber.open(io.BytesIO(content)) as pdf:
                    return "\n".join(page.extract_text() or "" for page in pdf.pages)

            if extension in self.IMAGE_EXTENSIONS:
                image = Image.open(io.BytesIO(content))
                return pytesseract.image_to_string(image, lang=OCR_LANGUAGES)

        except Exception as exc:
            log.warning("OCR failed for extension %s: %s", extension, exc)

        return ""

    # ── Amount extraction ──────────────────────────────────────────────

    def _extract_amount(self, text: str) -> Optional[float]:
        """
        Three-tier priority strategy:
        1. 'Bezahlter Betrag' / 'Paid' labels (what actually hit the bank — best match).
        2. 'Total' / 'Gesamtbetrag' labels, skipping 'Vorheriger' (historical) ones.
        3. Last currency-marked amount in the document (works for 2-column receipt layouts
           where the label text and amount appear in separate blocks).
        """
        if not text:
            return None

        paid = self._matches_with_filter(text, self.PAID_KEYWORD_PATTERN, group=1)
        if paid:
            return round(paid[-1], 2)

        totals = self._matches_with_filter(text, self.TOTAL_KEYWORD_PATTERN, group=1)
        if totals:
            return round(totals[-1], 2)

        all_currency = self._all_currency_amounts(text)
        if all_currency:
            return round(all_currency[-1], 2)

        return None

    def _matches_with_filter(self, text: str, pattern: re.Pattern, group: int) -> list[float]:
        """Returns parsed amounts for all matches, skipping those preceded by stale-amount words."""
        results: list[float] = []
        for match in pattern.finditer(text):
            # Check the 40 chars before the match for stale-amount words.
            context_before = text[max(0, match.start() - 40):match.start()].lower()
            if any(word in context_before for word in self.STALE_AMOUNT_PREFIXES):
                continue
            value = parse_number(match.group(group))
            if value is not None and self.AMOUNT_MIN < value < self.AMOUNT_MAX:
                results.append(value)
        return results

    def _all_currency_amounts(self, text: str) -> list[float]:
        """Extracts every currency-marked amount in document order."""
        values: list[float] = []
        for match in self.CURRENCY_AMOUNT_PATTERN.finditer(text):
            raw = match.group(1) or match.group(2)
            value = parse_number(raw)
            if value is not None and self.AMOUNT_MIN < value < self.AMOUNT_MAX:
                values.append(value)
        return values

    # ── Date extraction ────────────────────────────────────────────────

    def _extract_date(self, text: str) -> Optional[str]:
        if not text:
            return None

        # Numeric formats first (more reliable)
        for pattern in self.DATE_PATTERNS:
            match = pattern.search(text)
            if match:
                parsed = parse_date(match.group(1))
                if parsed:
                    return parsed

        # Fallback: German/English month names, e.g. "19. Februar 2026"
        return parse_date_with_month_name(text)


# ── Matching Service ──────────────────────────────────────────────────────

class MatchingService:
    """
    Pairs each transaction with its best-matching receipt.
    Uses a greedy strategy: each receipt can only be assigned once.

    For transactions with no confident match, a list of best-guess suggestions
    is provided so the user can manually assign the right one in the UI.
    """

    CONFIDENCE_SCORE = {"High": 3, "Medium": 2, "Low": 1, "None": 0}

    # How close an amount must be (relative) to still be suggested as a best guess.
    SUGGESTION_MAX_RELATIVE_DIFF = 0.20     # 20 %
    SUGGESTION_MAX_ABSOLUTE_DIFF = 50.0     # or €50, whichever is larger
    SUGGESTION_LIMIT = 5

    def match(self, transactions: list[BankTransaction],
              receipts: list[Receipt]) -> list[dict]:
        assigned_receipt_names: set[str] = set()
        results: list[dict] = []

        for txn in transactions:
            ranked = self._rank_candidates(txn, receipts, assigned_receipt_names)

            best_receipt, best_confidence = (
                (ranked[0][0], ranked[0][1]) if ranked else (None, "None")
            )

            if best_receipt is not None:
                assigned_receipt_names.add(best_receipt.fileName)
                # For matched rows: alternatives = other receipts that also matched
                alternatives = [r for r, _ in ranked[1:1 + MAX_ALTERNATIVES]]
            else:
                # For unmatched rows: alternatives = best-guess suggestions
                alternatives = self._suggest_for_unmatched(
                    txn, receipts, assigned_receipt_names
                )

            results.append({
                "transaction": asdict(txn),
                "matchedReceipt": asdict(best_receipt) if best_receipt else None,
                "confidence": best_confidence,
                "alternativeCandidates": [asdict(r) for r in alternatives],
            })

        return results

    # ── Confident matches ─────────────────────────────────────────────────

    def _rank_candidates(self, txn: BankTransaction,
                         receipts: list[Receipt],
                         assigned: set[str]) -> list[tuple[Receipt, str]]:
        scored = [
            (r, self._calculate_confidence(txn, r))
            for r in receipts if r.fileName not in assigned
        ]
        scored = [(r, c) for r, c in scored if c != "None"]
        scored.sort(key=lambda pair: self.CONFIDENCE_SCORE[pair[1]], reverse=True)
        return scored

    def _calculate_confidence(self, txn: BankTransaction, receipt: Receipt) -> str:
        if receipt.extractedAmount is None:
            return "None"

        tolerance = max(txn.amount * AMOUNT_TOLERANCE_PERCENT, AMOUNT_MIN_TOLERANCE)
        if abs(txn.amount - receipt.extractedAmount) > tolerance:
            return "None"

        if not receipt.extractedDate:
            return "Low"

        days_delta = abs(
            (datetime.fromisoformat(txn.date) - datetime.fromisoformat(receipt.extractedDate)).days
        )
        if days_delta <= DAYS_TOLERANCE_HIGH:
            return "High"
        if days_delta <= DAYS_TOLERANCE_MEDIUM:
            return "Medium"
        return "Low"

    # ── Best-guess suggestions for unmatched transactions ─────────────────

    def _suggest_for_unmatched(self, txn: BankTransaction,
                               receipts: list[Receipt],
                               assigned: set[str]) -> list[Receipt]:
        """
        Returns up to SUGGESTION_LIMIT receipts ranked by combined similarity score.
        Considers receipts whose amount is within a relaxed tolerance (20 % or €50).
        """
        available = [
            r for r in receipts
            if r.fileName not in assigned and r.extractedAmount is not None
        ]

        scored: list[tuple[float, Receipt]] = []
        for receipt in available:
            amount_diff = abs(txn.amount - receipt.extractedAmount)
            relative_diff = amount_diff / max(txn.amount, 1.0)

            # Skip receipts that are clearly unrelated
            if (relative_diff > self.SUGGESTION_MAX_RELATIVE_DIFF
                    and amount_diff > self.SUGGESTION_MAX_ABSOLUTE_DIFF):
                continue

            score = self._suggestion_score(txn, receipt, amount_diff)
            scored.append((score, receipt))

        # Lower score = better match
        scored.sort(key=lambda pair: pair[0])
        return [r for _, r in scored[:self.SUGGESTION_LIMIT]]

    @staticmethod
    def _suggestion_score(txn: BankTransaction, receipt: Receipt,
                          amount_diff: float) -> float:
        """
        Combined score: amount difference weighted heavily, date proximity as tiebreaker.
        """
        amount_component = amount_diff  # euros off
        if receipt.extractedDate:
            days_delta = abs(
                (datetime.fromisoformat(txn.date)
                 - datetime.fromisoformat(receipt.extractedDate)).days
            )
            date_component = min(days_delta, 365) * 0.05
        else:
            date_component = 5  # small penalty for missing date
        return amount_component + date_component


# ── Excel + ZIP Export ───────────────────────────────────────────────────

class ExportBuilder:
    """
    Produces a ZIP archive containing:
      - overview.xlsx — 4 sheets (Summary, Matched, Unmatched, Unused)
      - receipts/<yyyy-mm>/<renamed-file> — copies of matched receipts,
        renamed to `<date>_<amount>_<description>.<ext>`
      - receipts/_unused/<original-name> — receipts that weren't assigned

    The Excel file is structured so the accountant can:
      - Immediately see what needs follow-up (Unmatched sheet).
      - Drag renamed receipts into their bookkeeping software as evidence.
      - Spot duplicates or wrong-account receipts (Unused sheet).
    """

    # openpyxl styles (class-level so they're built once)
    _HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    _HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
    _HEADER_ALIGN = Alignment(horizontal="left", vertical="center")
    _CONF_FILLS = {
        "High":   PatternFill("solid", fgColor="DCFCE7"),
        "Medium": PatternFill("solid", fgColor="FEF3C7"),
        "Low":    PatternFill("solid", fgColor="FFEDD5"),
        "Manual": PatternFill("solid", fgColor="F3E8FF"),
    }
    _THIN_BORDER = Border(bottom=Side(style="thin", color="E2E8F0"))

    def __init__(self, store: "SessionStore"):
        self.store = store

    def build(self, session_id: str, results: list[dict]) -> tuple[bytes, str]:
        """Returns (zip_bytes, download_filename)."""
        session_dir = self.store.root / session_id
        if not session_dir.exists():
            raise ValueError("Session not found or expired.")

        matched, unmatched = self._partition_results(results)
        assigned_filenames = {r["matchedReceipt"]["fileName"] for r in matched}
        unused_filenames = self._find_unused(session_dir, assigned_filenames)

        # Build rename mapping: for each matched receipt → new monthly path
        rename_map = self._build_rename_map(matched)

        # Generate Excel file
        xlsx_bytes = self._build_excel(matched, unmatched, unused_filenames, rename_map)

        # Pack everything into a ZIP
        zip_bytes = self._build_zip(session_dir, xlsx_bytes, rename_map, unused_filenames)
        filename = self._download_filename(matched + unmatched)
        return zip_bytes, filename

    # ── Categorisation ────────────────────────────────────────────────────

    @staticmethod
    def _partition_results(results: list[dict]) -> tuple[list[dict], list[dict]]:
        matched, unmatched = [], []
        for r in results:
            (matched if r.get("matchedReceipt") else unmatched).append(r)
        return matched, unmatched

    def _find_unused(self, session_dir: Path, assigned: set[str]) -> list[str]:
        return sorted(
            f.name for f in session_dir.iterdir()
            if f.is_file() and f.name not in assigned
        )

    # ── Filename generation ───────────────────────────────────────────────

    def _build_rename_map(self, matched: list[dict]) -> dict[str, str]:
        """Maps original filename → new relative path (e.g. '2026-02/2026-02-19_380.80_GMAP.pdf')."""
        rename_map: dict[str, str] = {}
        used_paths: set[str] = set()

        for r in matched:
            original = r["matchedReceipt"]["fileName"]
            txn = r["transaction"]
            month = txn["date"][:7]  # YYYY-MM
            ext = Path(original).suffix or ""

            base_name = self._make_safe_filename(txn, ext)
            rel_path = f"{month}/{base_name}"
            rel_path = self._ensure_unique(rel_path, used_paths)
            used_paths.add(rel_path)
            rename_map[original] = rel_path

        return rename_map

    @staticmethod
    def _make_safe_filename(txn: dict, extension: str) -> str:
        """Builds `YYYY-MM-DD_AMOUNT_DESCRIPTION.ext` with sanitised parts."""
        date = txn["date"]
        amount = f"{txn['amount']:.2f}"
        raw_desc = txn.get("description", "")

        # Keep unicode letters/digits/hyphens; drop bullet chars, slashes, etc.
        safe = re.sub(r"[^\w\s-]", "", raw_desc, flags=re.UNICODE)
        safe = re.sub(r"\s+", "-", safe.strip())
        safe = safe[:40].rstrip("-") or "transaction"

        return f"{date}_{amount}_{safe}{extension}"

    @staticmethod
    def _ensure_unique(rel_path: str, used: set[str]) -> str:
        if rel_path not in used:
            return rel_path
        stem, ext = os.path.splitext(rel_path)
        counter = 2
        while f"{stem}_{counter}{ext}" in used:
            counter += 1
        return f"{stem}_{counter}{ext}"

    # ── Excel generation ──────────────────────────────────────────────────

    def _build_excel(self, matched: list[dict], unmatched: list[dict],
                     unused: list[str], rename_map: dict[str, str]) -> bytes:
        wb = openpyxl.Workbook()

        self._write_summary_sheet(wb.active, matched, unmatched, unused)
        self._write_matched_sheet(wb.create_sheet("Matched"), matched, rename_map)
        self._write_unmatched_sheet(wb.create_sheet("Unmatched"), unmatched)
        self._write_unused_sheet(wb.create_sheet("Unused Receipts"), unused)

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _write_summary_sheet(self, ws, matched: list[dict], unmatched: list[dict],
                             unused: list[str]) -> None:
        ws.title = "Summary"
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 20

        # Period derived from transaction dates
        all_results = matched + unmatched
        dates = [r["transaction"]["date"] for r in all_results if r["transaction"].get("date")]
        period = f"{min(dates)} to {max(dates)}" if dates else "—"

        # Counts by confidence and type
        conf_counts = {"High": 0, "Medium": 0, "Low": 0, "Manual": 0, "None": 0}
        total_debit = total_credit = 0.0
        for r in all_results:
            conf_counts[r.get("confidence", "None")] = conf_counts.get(r.get("confidence", "None"), 0) + 1
            t = r["transaction"]
            if t.get("type") == "Credit":
                total_credit += t["amount"]
            else:
                total_debit += t["amount"]

        # Write sections
        self._write_heading(ws, 1, "DocMatcher Export")
        ws.cell(row=2, column=1, value="Period").font = Font(bold=True)
        ws.cell(row=2, column=2, value=period)
        ws.cell(row=3, column=1, value="Generated").font = Font(bold=True)
        ws.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

        self._write_heading(ws, 5, "Transactions")
        rows = [
            ("Total", len(all_results)),
            ("Matched – high confidence", conf_counts["High"]),
            ("Matched – medium confidence", conf_counts["Medium"]),
            ("Matched – low confidence", conf_counts["Low"]),
            ("Manually assigned", conf_counts["Manual"]),
            ("Unmatched (follow up)", conf_counts["None"]),
        ]
        for i, (label, value) in enumerate(rows, start=6):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=value)

        self._write_heading(ws, 13, "Financial summary")
        ws.cell(row=14, column=1, value="Total expenses")
        c = ws.cell(row=14, column=2, value=total_debit)
        c.number_format = '#,##0.00 "EUR"'
        ws.cell(row=15, column=1, value="Total income")
        c = ws.cell(row=15, column=2, value=total_credit)
        c.number_format = '#,##0.00 "EUR"'

        self._write_heading(ws, 17, "Receipts")
        ws.cell(row=18, column=1, value="Matched to transactions").value = "Matched to transactions"
        ws.cell(row=18, column=2, value=len(matched))
        ws.cell(row=19, column=1, value="Unused (review for duplicates)")
        ws.cell(row=19, column=2, value=len(unused))

    def _write_matched_sheet(self, ws, matched: list[dict],
                             rename_map: dict[str, str]) -> None:
        headers = ["Date", "Description", "Amount", "Currency", "Type",
                   "Receipt File", "Confidence", "OCR Amount", "OCR Date"]
        self._write_header_row(ws, headers)

        for i, r in enumerate(matched, start=2):
            t = r["transaction"]
            mr = r["matchedReceipt"]
            row_data = [
                t["date"],
                t["description"],
                t["amount"],
                t["currency"],
                t["type"],
                rename_map.get(mr["fileName"], mr["fileName"]),
                r["confidence"],
                mr.get("extractedAmount"),
                mr.get("extractedDate") or "",
            ]
            for j, val in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                cell.border = self._THIN_BORDER
                if j in (3, 8) and val is not None and val != "":
                    cell.number_format = "#,##0.00"

            # Highlight confidence cell
            fill = self._CONF_FILLS.get(r["confidence"])
            if fill:
                ws.cell(row=i, column=7).fill = fill

        self._freeze_and_size(ws, [12, 40, 12, 10, 10, 42, 13, 12, 12])

    def _write_unmatched_sheet(self, ws, unmatched: list[dict]) -> None:
        headers = ["Date", "Description", "Amount", "Currency", "Type", "Notes"]
        self._write_header_row(ws, headers)

        for i, r in enumerate(unmatched, start=2):
            t = r["transaction"]
            row_data = [t["date"], t["description"], t["amount"],
                        t["currency"], t["type"], ""]
            for j, val in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=j, value=val)
                cell.border = self._THIN_BORDER
                if j == 3:
                    cell.number_format = "#,##0.00"

        self._freeze_and_size(ws, [12, 40, 12, 10, 10, 30])

    def _write_unused_sheet(self, ws, unused_filenames: list[str]) -> None:
        headers = ["File Name", "Notes"]
        self._write_header_row(ws, headers)

        for i, fn in enumerate(unused_filenames, start=2):
            ws.cell(row=i, column=1, value=fn).border = self._THIN_BORDER
            ws.cell(row=i, column=2, value="").border = self._THIN_BORDER

        self._freeze_and_size(ws, [50, 30])

    def _write_heading(self, ws, row: int, text: str) -> None:
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=12, color="1E3A8A")

    def _write_header_row(self, ws, headers: list[str]) -> None:
        for col, name in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = self._HEADER_FONT
            cell.fill = self._HEADER_FILL
            cell.alignment = self._HEADER_ALIGN

    def _freeze_and_size(self, ws, widths: list[int]) -> None:
        ws.freeze_panes = "A2"
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 22

    # ── ZIP packaging ─────────────────────────────────────────────────────

    def _build_zip(self, session_dir: Path, xlsx_bytes: bytes,
                   rename_map: dict[str, str], unused: list[str]) -> bytes:
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("overview.xlsx", xlsx_bytes)

            for original, new_path in rename_map.items():
                src = session_dir / original
                if src.exists():
                    zf.write(src, f"receipts/{new_path}")

            for fn in unused:
                src = session_dir / fn
                if src.exists():
                    zf.write(src, f"receipts/_unused/{fn}")

        return zip_buffer.getvalue()

    @staticmethod
    def _download_filename(results: list[dict]) -> str:
        dates = [r["transaction"]["date"] for r in results
                 if r["transaction"].get("date")]
        if not dates:
            return "docmatcher-export.zip"
        earliest = min(dates)[:7]  # YYYY-MM
        latest = max(dates)[:7]
        suffix = earliest if earliest == latest else f"{earliest}_to_{latest}"
        return f"docmatcher-export_{suffix}.zip"


# ── Flask App ─────────────────────────────────────────────────────────────

app = Flask(__name__, static_folder="static")

csv_parser = CsvParser()
ocr_service = OcrService()
matching_service = MatchingService()


# ── Session Storage ──────────────────────────────────────────────────────

class SessionStore:
    """Stores uploaded receipt files on disk for later preview."""

    def __init__(self, root: Path, ttl_hours: int):
        self.root = root
        self.ttl_seconds = ttl_hours * 3600
        self.root.mkdir(parents=True, exist_ok=True)

    def create_session(self) -> tuple[str, Path]:
        session_id = uuid.uuid4().hex
        session_dir = self.root / session_id
        session_dir.mkdir(parents=True, exist_ok=True)
        return session_id, session_dir

    def save_file(self, session_dir: Path, filename: str, content: bytes) -> str:
        """Saves a file with a sanitised name. Returns the stored name."""
        safe_name = secure_filename(filename) or f"file_{uuid.uuid4().hex[:8]}"
        (session_dir / safe_name).write_bytes(content)
        return safe_name

    def get_file_path(self, session_id: str, filename: str) -> Optional[Path]:
        """Resolves a session file path, preventing path traversal."""
        safe_name = secure_filename(filename)
        if not safe_name:
            return None
        try:
            session_dir = (self.root / session_id).resolve()
            if self.root not in session_dir.parents and session_dir != self.root:
                return None
            filepath = (session_dir / safe_name).resolve()
            if session_dir not in filepath.parents:
                return None
            return filepath if filepath.exists() else None
        except (OSError, ValueError):
            return None

    def cleanup_expired(self) -> int:
        """Removes session directories older than TTL. Returns count removed."""
        if not self.root.exists():
            return 0
        cutoff = time.time() - self.ttl_seconds
        removed = 0
        for session_dir in self.root.iterdir():
            if session_dir.is_dir() and session_dir.stat().st_mtime < cutoff:
                shutil.rmtree(session_dir, ignore_errors=True)
                removed += 1
        return removed


session_store = SessionStore(UPLOAD_DIR, SESSION_TTL_HOURS)
export_builder = ExportBuilder(session_store)


@app.route("/")
def index():
    return send_from_directory("static", "index.html")


def _ndjson(obj: dict) -> str:
    return json.dumps(obj) + "\n"


@app.post("/api/match")
def api_match():
    """
    Streams progress events as newline-delimited JSON (NDJSON) while OCR runs,
    so the frontend can show a live progress bar. The final line is either
    {"type":"result", ...} on success or {"type":"error", ...} on failure.
    """
    csv_file = request.files.get("csvFile")
    receipt_files = request.files.getlist("receiptFiles")

    if not csv_file:
        return jsonify({"error": "csvFile is required"}), 400
    if not receipt_files:
        return jsonify({"error": "At least one receipt file is required"}), 400

    # Read uploads fully inside the request context — the streaming generator
    # below runs after the request has been consumed, so `request.files` would
    # no longer be accessible there.
    csv_bytes = csv_file.read()
    uploaded: list[tuple[str, bytes]] = [
        (f.filename, f.read()) for f in receipt_files
    ]

    def stream():
        try:
            yield _ndjson({"type": "progress", "stage": "parsing_csv"})
            log.info("Parsing CSV: %s", csv_file.filename)
            transactions = csv_parser.parse(csv_bytes)

            if not transactions:
                yield _ndjson({
                    "type": "error",
                    "error": "No valid transactions found in CSV. "
                             "Check that column headers include a Date and "
                             "Amount (or Debit/Credit).",
                })
                return

            session_id, session_dir = session_store.create_session()
            total = len(uploaded)
            log.info("Parsed %d transactions. Running OCR on %d receipts (session %s)…",
                     len(transactions), total, session_id[:8])

            yield _ndjson({
                "type": "progress", "stage": "ocr_start",
                "transactions": len(transactions), "total": total,
            })

            receipts = []
            for idx, (filename, content) in enumerate(uploaded, start=1):
                stored_name = session_store.save_file(session_dir, filename, content)
                receipts.append(ocr_service.extract(content, stored_name))
                yield _ndjson({
                    "type": "progress", "stage": "ocr",
                    "current": idx, "total": total, "file": filename,
                })

            yield _ndjson({"type": "progress", "stage": "matching"})
            log.info("OCR complete. Running matching…")
            results = matching_service.match(transactions, receipts)

            matched_count = sum(1 for r in results if r["confidence"] != "None")
            log.info("Done. Matched %d of %d transactions.", matched_count, len(transactions))

            removed = session_store.cleanup_expired()
            if removed:
                log.info("Cleaned up %d expired sessions.", removed)

            yield _ndjson({
                "type": "result",
                "sessionId": session_id,
                "results": results,
            })

        except Exception as exc:
            log.error("Request failed: %s", exc)
            traceback.print_exc()
            yield _ndjson({"type": "error", "error": str(exc)})

    # X-Accel-Buffering disables proxy buffering (nginx) so events stream live.
    return Response(
        stream(),
        mimetype="application/x-ndjson",
        headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"},
    )


@app.get("/api/receipt/<session_id>/<path:filename>")
def api_receipt(session_id: str, filename: str):
    """Serves a receipt file from a specific session for in-browser preview."""
    filepath = session_store.get_file_path(session_id, filename)
    if filepath is None:
        abort(404)
    # Inline disposition so PDFs/images render in the browser instead of downloading
    return send_from_directory(
        str(filepath.parent),
        filepath.name,
        as_attachment=False,
    )


@app.post("/api/export")
def api_export():
    """
    Generates a ZIP archive with renamed receipts + overview.xlsx.
    The frontend sends the current results including any manual assignments.
    """
    try:
        payload = request.get_json(silent=True) or {}
        session_id = payload.get("sessionId")
        results = payload.get("results", [])

        if not session_id or not results:
            return jsonify({"error": "sessionId and results are required"}), 400

        log.info("Building export for session %s (%d transactions)…",
                 session_id[:8], len(results))

        zip_bytes, filename = export_builder.build(session_id, results)

        log.info("Export ready: %s (%.1f KB)", filename, len(zip_bytes) / 1024)
        return send_file(
            BytesIO(zip_bytes),
            mimetype="application/zip",
            as_attachment=True,
            download_name=filename,
        )

    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        log.error("Export failed: %s", exc)
        traceback.print_exc()
        return jsonify({"error": str(exc)}), 500


if __name__ == "__main__":
    # Verify Tesseract is available on startup
    try:
        version = pytesseract.get_tesseract_version()
        log.info("Tesseract %s ready. Languages: %s", version, OCR_LANGUAGES)
    except Exception as exc:
        log.warning("Tesseract not found: %s", exc)
        log.warning("Install it first:")
        log.warning("  macOS:  brew install tesseract")
        log.warning("  Ubuntu: sudo apt install tesseract-ocr")
        log.warning("  Windows: https://github.com/UB-Mannheim/tesseract/wiki")

    log.info("Starting DocMatcher on http://localhost:%d", PORT)
    app.run(host="0.0.0.0", port=PORT, debug=False)
